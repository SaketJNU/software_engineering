import csv
#import json
import openpyxl
import os





def list_all_files(folder_name):
    file_list = os.listdir(folder_name)
    return file_list

def read_csv(csv_folder_name):
    file_list = list_all_files(csv_folder_name)
    print("It has the file with name {}".format(file_list))

    for file in file_list:
        print("file_name is {}".format(file))
        exampleFile = open(csv_folder_name + "/" + file)
        file_data1 = csv.reader(exampleFile)
        file_data = list(file_data1)
    return file_data



def filter_date_data(date, csv_file):
    one_date_data = []
    for row in csv_file:
        if date in row[0]:
            one_date_data.append(row[2])
    return one_date_data


def find_date_list(csv_file):
    date_list = []
    for row in csv_file:
        new_date = row[0].split(" ")[0]
        if new_date != "timestamp":
            if new_date not in date_list:
                date_list.append(new_date)
                print("A new date {} has been added in the date list".format(new_date))
    return date_list


def save_day_data(one_date_data, date, xlsx_folder_name):
    file_name = xlsx_folder_name + "/" + date + ".xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = "Price_Data"
    i = 1
    for data in one_date_data:
        sheet.cell(row = i, column = 1).value = data
        i = i + 1
    wb.save(file_name)

    #######################################

    #file_name1 = xlsx_folder_name + "1/" + date + "xlsx"
    #wb1 = openpyxl.Workbook()
    #sheet = wb1.get_active_sheet()
    #sheet.title = "Price_Data"
    #sheet['A1'] = one_date_data
    #wb1.save(file_name1)
    return


if __name__ == "__main__":
    csv_folder_name = "raw_data"
    json_folder_name = "processed_data"
    xlsx_folder = "excel_data"
    csv_file = read_csv(csv_folder_name)
    #print("CSV File data is: {}", csv_file)
    #for row in csv_file:
    #    print(row)
    date_list = find_date_list(csv_file)
    #print(" ---------------------------------------------------")
    #print("All dates are = {}".format(date_list))
    for date in date_list:
        one_date_data = filter_date_data(date, csv_file)
        #print("Data for date {} is {}".format(date, one_date_data))
        save_day_data(one_date_data, date, xlsx_folder)
